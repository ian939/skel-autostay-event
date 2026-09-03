# 검증된 지점_좌표.json에서 이벤트 참여 15개 지점만 골라 assets/js/stores.js를 생성한다
"""
데이터 출처와 근거:
  - 좌표: <분석폴더>/output/지점_좌표.json
    카카오 로컬 API로 29/29 주소 정확매칭, 전 지점 정밀도 '높음', 시군구일치 True.
    (근거: 분석폴더 context-notes.md F6)
  - 참여 여부: <분석폴더>/오토스테이_지점별_반경분석(분석)_참여매장표시_260902.xlsx
    시트 '지점별_반경분석'에서 A열 배경색이 FFFFFF00(노란색)인 행 = 이벤트 참여 지점.

손으로 옮겨 적지 않는 이유: 좌표 15개 x 소수점 6자리를 수동 전사하면 오타가 난다.
16번째 지점이 추가되면 엑셀에 음영만 칠하고 이 스크립트를 다시 돌리면 된다.
"""
import json
import os
import sys
import unicodedata

# 분석 폴더(원본 데이터가 있는 곳). 이 리포와 물리적으로 분리되어 있다.
SRC_DIR = r"c:\Users\LEGION\Desktop\마케팅셀\01. Idea\14. 제휴\오토스테이"
COORD_JSON = os.path.join(SRC_DIR, "output", "지점_좌표.json")
OUT_JS = os.path.join(os.path.dirname(__file__), "..", "assets", "js", "stores.js")

YELLOW = "FFFFFF00"

# 지역 -> 필터 칩 그룹. 조건문을 코드 곳곳에 흩뿌리지 않기 위해 데이터로 둔다.
REGION_GROUP = {
    "고양": "goyang", "파주": "goyang",
    "서울": "seoul",  "광명": "seoul",
    "양주": "north",  "하남": "north",
    "용인": "south",  "평택": "south", "안성": "south",
    "대구": "daegu",  "구미": "daegu",
}


def find_flag_xlsx():
    # Windows에서 한글 파일명이 NFD(자모 분리)로 저장되어 있어 조합형 패턴 매칭이 실패한다.
    # 그래서 디렉터리를 훑으며 NFC 정규화 후 비교한다.
    want = unicodedata.normalize("NFC", "참여매장표시")
    hits = []
    for fn in os.listdir(SRC_DIR):
        if not fn.lower().endswith(".xlsx"):
            continue
        if want in unicodedata.normalize("NFC", fn):
            hits.append(os.path.join(SRC_DIR, fn))
    if not hits:
        sys.exit(f"[FAIL] 참여매장표시 엑셀을 찾을 수 없음: {SRC_DIR}")
    return max(hits, key=os.path.getmtime)


def read_participating_names(path):
    """A열이 노란색으로 음영된 행의 지점명을 모은다."""
    import openpyxl
    wb_v = openpyxl.load_workbook(path, data_only=True)
    wb_s = openpyxl.load_workbook(path)
    ws_v, ws_s = wb_v["지점별_반경분석"], wb_s["지점별_반경분석"]

    names = []
    for r in range(2, ws_v.max_row + 1):
        fill = ws_s.cell(r, 1).fill
        rgb = str(fill.start_color.rgb) if fill and fill.start_color else ""
        if rgb == YELLOW:
            names.append(ws_v.cell(r, 4).value)
    return names


def main():
    xlsx = find_flag_xlsx()
    names = read_participating_names(xlsx)
    print(f"[1] 참여 지점(음영): {len(names)}개  <- {os.path.basename(xlsx)}")

    with open(COORD_JSON, encoding="utf-8") as f:
        coords = json.load(f)
    by_name = {c["지점명"]: c for c in coords}
    print(f"[2] 좌표 원본: {len(coords)}개")

    missing = [n for n in names if n not in by_name]
    if missing:
        sys.exit(f"[FAIL] 좌표를 찾을 수 없는 지점: {missing}")

    records = []
    for i, n in enumerate(sorted(names, key=lambda x: by_name[x]["no"]), start=1):
        c = by_name[n]
        region = c["지역"]
        if region not in REGION_GROUP:
            sys.exit(f"[FAIL] REGION_GROUP에 없는 지역: '{region}' ({n}) -> 매핑을 추가하라")
        # 좌표는 소수점 6자리 = 약 11cm. 지점 찾기 용도에 충분하고 파일이 작아진다.
        records.append({
            "id": f"as-{i:02d}",
            "no": c["no"],
            "name": n,
            "addr": c["주소"],
            "hours": c["영업시간"],
            "type": c["구분"],
            "region": region,
            "zone": c["권역"],
            "group": REGION_GROUP[region],
            "lat": round(c["위도"], 6),
            "lng": round(c["경도"], 6),
        })

    # 집계 검증: 계획서에 기록된 값과 일치해야 한다.
    zones, types, groups = {}, {}, {}
    for r in records:
        zones[r["zone"]] = zones.get(r["zone"], 0) + 1
        types[r["type"]] = types.get(r["type"], 0) + 1
        groups[r["group"]] = groups.get(r["group"], 0) + 1

    print(f"[3] 권역: {zones}")
    print(f"[4] 구분: {types}")
    print(f"[5] 그룹: {groups}")

    expect = {"zones": {"수도권": 12, "비수도권": 3}, "types": {"직영": 7, "가맹": 8}}
    if zones != expect["zones"] or types != expect["types"]:
        print(f"[WARN] 집계가 계획서 기준({expect})과 다르다. 참여 지점이 변경되었는지 확인하라.")
    if sum(groups.values()) != len(records):
        sys.exit("[FAIL] 그룹 합계 불일치")

    body = ",\n".join(
        "  { id: %s, no: %d, name: %s, addr: %s, hours: %s,\n"
        "    type: %s, region: %s, zone: %s, group: %s, lat: %s, lng: %s }" % (
            json.dumps(r["id"], ensure_ascii=False), r["no"],
            json.dumps(r["name"], ensure_ascii=False),
            json.dumps(r["addr"], ensure_ascii=False),
            json.dumps(r["hours"], ensure_ascii=False),
            json.dumps(r["type"], ensure_ascii=False),
            json.dumps(r["region"], ensure_ascii=False),
            json.dumps(r["zone"], ensure_ascii=False),
            json.dumps(r["group"], ensure_ascii=False),
            r["lat"], r["lng"],
        )
        for r in records
    )

    js = (
        "// 이벤트 참여 오토스테이 지점 데이터 (tools/build_stores.py 자동생성 - 직접 수정 금지)\n"
        "// 좌표 출처: 카카오 로컬 API 지오코딩, 주소 정확매칭 + 시군구 검증 통과\n"
        f"// 생성 기준 파일: {os.path.basename(xlsx)}\n"
        f"// 지점 수: {len(records)}개 (수도권 {zones.get('수도권', 0)} / 비수도권 {zones.get('비수도권', 0)}"
        f" · 직영 {types.get('직영', 0)} / 가맹 {types.get('가맹', 0)})\n"
        "\n"
        "const STORES = [\n" + body + "\n];\n"
    )

    out = os.path.abspath(OUT_JS)
    with open(out, "w", encoding="utf-8", newline="\n") as f:
        f.write(js)
    print(f"[6] 생성 완료: {out} ({len(records)}개, {len(js):,} bytes)")


if __name__ == "__main__":
    main()
